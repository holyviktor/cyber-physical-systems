from datetime import datetime

import openpyxl
import pandas as pd


class Database:
    def __init__(self, path):
        self.path = path
        wb_obj = openpyxl.load_workbook(path)
        self.sheet_obj = wb_obj.active
        self.maxRows = self.sheet_obj.max_row
        self.maxColumns = self.sheet_obj.max_column

    def print_data(self):
        columns = []
        print("\nValue of first column")
        for i in range(2, self.maxRows + 1):
            cell_obj = self.sheet_obj.cell(row=i, column=4)
            columns.append(cell_obj.value)
        print(columns)

        print("\nValue of first row")
        for i in range(1, self.maxColumns + 1):
            cell_obj = self.sheet_obj.cell(row=2, column=i)
            print(cell_obj.value, end=" ")

    def check_time(self):
        print("\nChecking time...")
        for i in range(2, self.maxRows):
            time1 = self.sheet_obj.cell(row=i, column=2)
            cell_obj1 = datetime.strptime(str(time1.value), "%H:%M:%S")
            cell_obj2 = datetime.strptime(str(self.sheet_obj.cell(row=i + 1, column=2).value), "%H:%M:%S")
            # print(type(str(cell_obj1.value)))
            timedelta = cell_obj2 - cell_obj1
            timedeltaneed1 = datetime.strptime("00:30:00", "%H:%M:%S") - datetime.strptime("00:00:00", "%H:%M:%S")
            timedeltaneed2 = datetime.strptime("00:00:00", "%H:%M:%S") - datetime.strptime("23:30:00", "%H:%M:%S")
            # print(timedelta)
            # print(timedeltaneed2)
            # print(cell_obj1)
            if timedelta != timedeltaneed1 and timedelta != timedeltaneed2:
                print("Time error", time1)

    def check_speed(self):
        list_speed = []
        ind_speed = False
        print("\nChecking speed...")
        for i in range(2, self.maxRows + 1):
            speed = self.sheet_obj.cell(row=i, column=5)
            list_speed.append(speed.value)
            if speed.value is None:
                print("Speed error", speed)
                ind_speed = True
        if ind_speed:
            pass

    def check_direction(self):
        print("\nChecking speed...")
        for i in range(2, self.maxRows + 1):
            direction = self.sheet_obj.cell(row=i, column=4)
            if direction.value is None:
                if not self.sheet_obj.cell(row=i, column=5).value:
                    # Штиль
                    pass
                else:
                    #Перемінний
                    pass

    def check_temperature(self):
        list_temperature = []
        ind_temperature = False
        print("\nChecking temperature...")
        for i in range(2, self.maxRows + 1):
            temperature = self.sheet_obj.cell(row=i, column=3)
            list_temperature.append(temperature.value)
            if temperature.value is None:
                print("Speed error", temperature)
                ind_temperature = True
        if ind_temperature:
            pass


def interpolate(interpolate_list, type_interpolation):
    a = pd.Series(interpolate_list)
    #a.interpolate()
    #a.interpolate(method="pad", limit=2)
    fixed_list = a.interpolate(method="polynomial", order=1)
    return fixed_list


def main():
    database = Database("2012-1.xlsx")
    database.print_data()
    database.check_speed()
    for i in interpolate([0, 1, None, 3, 4, 5, 7], ""):
        print(i)


if __name__ == '__main__':
    main()
